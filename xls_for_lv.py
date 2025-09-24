import pythoncom

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.drawing.image import Image
from math import ceil
from settings import ensure_valid_path
import os
from datetime import datetime

import win32com.client

from PyQt6 import QtWidgets, QtCore

def show_error(text: str, title="Ошибка"):
    message = QtWidgets.QMessageBox()
    message.setWindowFlag(QtCore.Qt.WindowType.FramelessWindowHint)
    message.setText(text)
    message.setIcon(QtWidgets.QMessageBox.Icon.Warning)
    message.setWindowTitle(title)
    message.setStyleSheet('''
        background-color: rgb(35, 40, 49);
        color: white;
        font-size: 12pt;
        font-weight: 700;
        font-family: RussoOne-Regular;
        border: 2px solid gray;
        border-radius: 10px;
    ''')
    message.exec()


def create_tax_lv(lst_for_csv: list, current_inv: str, rate_current: str) -> None:


    wb = Workbook()
    wb.guess_type = True
    ws = wb.active
    ws.title = 'ITMS - Calculation of duties'

    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.merge_cells('A1:E1')


    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'] = 'ITMS - Calculation of duties'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)

    ws.column_dimensions['A'].width = 14
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A3'].font = Font(name='Arial', size=10, bold=True)
    ws['A3'] = 'Commodity code'
    ws['A3'].border = thin_border

    ws.column_dimensions['B'].width = 10
    ws['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['B3'].font = Font(name='Arial', size=10, bold=True)
    ws['B3'] = 'Duty rate / PVN 21%'
    ws['B3'].border = thin_border

    ws.column_dimensions['C'].width = 8
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['C3'].font = Font(name='Arial', size=10, bold=True)
    ws['C3'] = 'Weight, KG'
    ws['C3'].border = thin_border

    ws.column_dimensions['D'].width = 14
    ws['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['D3'].font = Font(name='Arial', size=10, bold=True)
    ws['D3'] = 'Customs value, ' + current_inv
    ws['D3'].border = thin_border

    ws.column_dimensions['E'].width = 12
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E3'].font = Font(name='Arial', size=10, bold=True)
    ws['E3'] = 'Duty amount, EUR'
    ws['E3'].border = thin_border

    for i in range(0, len(lst_for_csv)):


        ws[f'A{i+4}'] = lst_for_csv[i][0]
        ws[f'A{i+4}'].border = thin_border

        if isinstance(lst_for_csv[i][1], float):
            ws[f'B{i+4}'] = lst_for_csv[i][1]
            ws[f'B{i+4}'].border = thin_border
        else:
            ws[f'B{i+4}'] = lst_for_csv[i][1]
            ws[f'B{i+4}'].border = thin_border
            ws[f'B{i+4}'].number_format = BUILTIN_FORMATS[10]
            ws[f'B{i+4}'].alignment = Alignment(horizontal='right')



        ws[f'C{i+4}'] = lst_for_csv[i][2]
        ws[f'C{i+4}'].border = thin_border

        ws[f'D{i+4}'] = lst_for_csv[i][3]
        ws[f'D{i+4}'].border = thin_border

        if lst_for_csv[i][4] != 'ERROR':


            ws[f'E{i+4}'].border = thin_border
            ws[f'E{i+4}'].number_format = BUILTIN_FORMATS[2]  # Установка числового формата
            ws[f'E{i+4}'] = float(lst_for_csv[i][4])
        else:
            ws[f'E{i+4}'] = f'=(((D{i+4}*B{i+4})+D{i+4})*0.21+(D{i+4}*B{i+4}))*{rate_current}'
            ws[f'E{i+4}'].number_format = BUILTIN_FORMATS[2]
            ws[f'E{i+4}'].border = thin_border



    ws['D' + str(len(lst_for_csv) + 5)] = 'Total: '
    ws['D' + str(len(lst_for_csv) + 5)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['D' + str(len(lst_for_csv) + 5)].font = Font(name='Arial', size=14, bold=True)

    ws['E' + str(len(lst_for_csv) + 5)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E' + str(len(lst_for_csv) + 5)].font = Font(name='Arial', size=14, bold=True)
    ws['E' + str(len(lst_for_csv) + 5)] = f'=SUM(E4:E{len(lst_for_csv) + 4})'
    ws['E' + str(len(lst_for_csv) + 5)].number_format = '#,##0.00" €"'
     # Папка для сохранения
    folder_path = ensure_valid_path("calc_for_lv")

    # Название файла (без расширения)
    now = datetime.now()
    base_name = f"TAX_{now.strftime('%d-%m-%Y_%H-%M')}.csv"
    #
    # Полный путь для Excel
    excel_path = os.path.join(folder_path, base_name + ".xlsx")
    pdf_path = os.path.join(folder_path, base_name + ".pdf")

    if os.path.exists(pdf_path):
        try:
            os.remove(pdf_path)
        except PermissionError:
            show_error("PDF-файл уже открыт в другом приложении.\nПожалуйста, закройте его и попробуйте снова.")
            return
        except Exception as e:
            show_error(f"Не удалось удалить старый PDF: {e}")
            return
    try:
        ws.page_setup.scale = 75
        wb.save(excel_path)
        del wb
    except PermissionError:
        show_error("Файл уже открыт. Пожалуйста, закройте его и повторите попытку.")
        return
    except Exception as e:
        show_error(f"Ошибка при сохранении Excel: {e}")
        return

        # Сохраняем как PDF
    pythoncom.CoInitialize()
    excel_app = win32com.client.DispatchEx("Excel.Application")
    excel_app.Visible = 0

    try:
        wb1 = excel_app.Workbooks.Open(excel_path)
        wb1.ExportAsFixedFormat(0, pdf_path)
        wb1.Close(SaveChanges=False)
    except Exception as e:
        show_error(f"Ошибка при экспорте в PDF: {e}")
        return
    finally:
        if 'excel_app' in locals():
            excel_app.Quit()  # Закрывает ТОЛЬКО созданный экземпляр
            del excel_app
