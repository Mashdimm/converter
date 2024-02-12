import pythoncom
import os


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.drawing.image import Image
from math import ceil

import win32com.client


def create_payment_vet(dannye_payment: dict) -> None:


    wb = Workbook()
    ws = wb.active
    ws.title = 'Payment'
    logo = Image('img.png')
    logo.height = 160
    logo.width = 155
    thins = Side(border_style="thin", color="000000")
    doubautole = Side(border_style="dashDot", color="ff0000")


    ws.merge_cells('B1:M1')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['B1'] = 'Dokumento, patvirtinančio išankstinį mokėjimą už Valstybinės maisto ir veterinarijos tarnybos'

    ws.merge_cells('B2:M2')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['B2'] = 'kontroliuojamų prekių siuntos valstybinę veterinarinę/ maisto kontrolę, formos pavyzdys'

    top_left_cell = ws['D3']
    top_left_cell.value = dannye_payment['dte_plat']

    top_left_cell.border = Border(bottom=thins)
    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('D3:E3')

    ws['G3'] = '№'
    ws['G3'].alignment = Alignment(horizontal='center')

    top_left_cell = ws['H3']
    top_left_cell.value = dannye_payment['platezhka']

    top_left_cell.border = Border(bottom=thins)
    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('H3:I3')

    ws.merge_cells('D4:E4')
    ws['D4'].alignment = Alignment(horizontal='center')
    ws['D4'] = 'Data'

    ws.merge_cells('B6:M6')
    ws['B6'].alignment = Alignment(horizontal='center')
    ws[
        'B6'] = 'Aš, žemiau pasirašęs MB Alesta LT įgaliotas asmuo, patvirtinu, kad už žemiau  nurodytą valstybinei veterinarinei / '
    ws['B6'].font = Font(size=8)

    ws.merge_cells('B7:M7')
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'] = 'maisto kontrolei pateikiamą siuntą yra iš anksto sumokėta valstybės rinkliava:'
    ws['B7'].font = Font(size=8)

    top_left_cell = ws['B9']
    top_left_cell.value = '1. Išankstinio mokėjimo duomenys'
    top_left_cell.font = Font(size=8)

    top_left_cell.border = Border(top=thins, left=thins)
    top_left_cell.alignment = Alignment(horizontal='left')
    ws.merge_cells('B9:D9')

    top_left_cell = ws['E9']
    top_left_cell.value = dannye_payment['dte_poruch']
    top_left_cell.font = Font(underline='single')

    top_left_cell.border = Border(top=thins, left=thins)
    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('E9:G9')

    top_left_cell = ws['H9']
    top_left_cell.value = dannye_payment['plat_poruch']
    top_left_cell.font = Font(underline='single')

    top_left_cell.border = Border(top=thins, left=thins)
    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('H9:J9')

    top_left_cell = ws['K9']
    top_left_cell.value = len(dannye_payment['chedp']) * 50
    top_left_cell.font = Font(underline='single')

    top_left_cell.border = Border(top=thins, left=thins, right=thins)
    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('K9:M9')

    top_left_cell = ws['B10']
    top_left_cell.font = Font(underline='single')

    top_left_cell.border = Border(left=thins, bottom=thins)
    ws.merge_cells('B10:D10')

    top_left_cell = ws['E10']
    top_left_cell.value = '(pavedimo data)'
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')

    top_left_cell.border = Border(left=thins, bottom=thins)
    ws.merge_cells('E10:G10')

    top_left_cell = ws['H10']
    top_left_cell.value = '(pavedimo numeris)'
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')

    top_left_cell.border = Border(left=thins, bottom=thins)
    ws.merge_cells('H10:J10')

    top_left_cell = ws['K10']
    top_left_cell.value = '(suma EUR)'
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')

    top_left_cell.border = Border(left=thins, bottom=thins, right=thins)
    ws.merge_cells('K10:M10')

    top_left_cell = ws['B11']
    top_left_cell.value = '2. Transporto priemonė:'
    top_left_cell.font = Font(size=8)

    top_left_cell.border = Border(left=thins)
    top_left_cell.alignment = Alignment(horizontal='left')
    ws.merge_cells('B11:C11')

    top_left_cell = ws['E11']
    top_left_cell.value = dannye_payment['numb_truck']
    top_left_cell.font = Font(underline='single')

    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('E11:J11')

    ws['M11'].border = Border(right=thins)

    top_left_cell = ws['B12']
    top_left_cell.value = '(automobilio ir priekabos valstybinis numeris)'
    top_left_cell.font = Font(size=8)

    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.border = Border(left=thins, right=thins, bottom=thins)
    ws.merge_cells('B12:M12')

    top_left_cell = ws['B13']
    top_left_cell.value = '3. Siuntą lydinčio veterinarijos sertifikato ar kito veterinarinio / saugą ir kokybę patvirtinančio dokumento numeris ir data: '
    top_left_cell.font = Font(size=8)

    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.border = Border(left=thins, right=thins)
    ws.merge_cells('B13:M13')

    top_left_cell = ws['B14']

    top_left_cell.value = ', '.join(dannye_payment['svid']) + ' ' + dannye_payment['dte_svid']

    top_left_cell.alignment = Alignment(horizontal='center', wrapText=True)
    top_left_cell.font = Font(underline='single')
    top_left_cell.border = Border(left=thins, right=thins, bottom=thins)
    if len(', '.join(dannye_payment['svid']) + ' ' + dannye_payment['dte_svid']) > 100:
        ws.row_dimensions[14].height = ceil(
            len(', '.join(dannye_payment['svid']) + ' ' + dannye_payment['dte_svid']) / 100) * 14.4
    ws.merge_cells('B14:M14')

    top_left_cell = ws['B15']
    top_left_cell.value = '4. Pasienio veterinarijos posto, pro kurį įvežama siunta  į Lietuvos Respublikos teritoriją, pavadinimas'
    top_left_cell.font = Font(size=8)

    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.border = Border(left=thins, right=thins)
    ws.merge_cells('B15:M15')

    top_left_cell = ws['B16']
    top_left_cell.value = dannye_payment['pereh']
    top_left_cell.font = Font(underline='single')

    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.border = Border(left=thins, right=thins, bottom=thins)
    ws.merge_cells('B16:M16')

    top_left_cell = ws['B17']
    top_left_cell.value = '5. MB Alesta LT įgaliotas asmuo: '
    top_left_cell.font = Font(size=8)

    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.border = Border(left=thins, right=thins)
    ws.merge_cells('B17:M17')

    top_left_cell = ws['B18']
    top_left_cell.value = '            Direktorius               '
    top_left_cell.font = Font(underline='single')

    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.border = Border(left=thins)
    ws.merge_cells('B18:E18')

    top_left_cell = ws['F18']
    top_left_cell.value = '___________________'
    # top_left_cell.font = Font(underline='single')

    top_left_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('F18:I18')

    top_left_cell = ws['J18']
    top_left_cell.value = '         Robert Žinis          '
    top_left_cell.font = Font(underline='single')

    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.border = Border(right=thins)
    ws.merge_cells('J18:M18')

    ws.add_image(logo, 'G17')

    top_left_cell = ws['B19']
    top_left_cell.value = '(pareigos)'
    top_left_cell.font = Font(size=8)

    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.border = Border(left=thins)
    ws.merge_cells('B19:E19')

    top_left_cell = ws['F19']
    top_left_cell.value = '(parašas)'

    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    ws.merge_cells('F19:I19')

    top_left_cell = ws['J19']
    top_left_cell.value = '(vardas, pavardė)'
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.border = Border(right=thins)
    ws.merge_cells('J19:M19')

    ws['B20'].border = Border(left=thins)
    ws['M20'].border = Border(right=thins)

    top_left_cell = ws['B21']
    top_left_cell.value = 'Antspaudas'
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.border = Border(left=thins)
    ws.merge_cells('B21:E21')

    ws['M21'].border = Border(right=thins)

    top_left_cell = ws['B22']
    top_left_cell.value = '6. Bendrojo sveikatos įvežimo dokumento: '
    top_left_cell.font = Font(size=8)
    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.border = Border(top=thins, left=thins, right=thins)
    ws.merge_cells('B22:M22')

    top_left_cell = ws['B23']

    top_left_cell.value = ', '.join(dannye_payment['chedp']) + ' ' + dannye_payment['dte_ched']

    top_left_cell.alignment = Alignment(horizontal='center', wrapText=True)
    top_left_cell.font = Font(underline='single')
    top_left_cell.border = Border(left=thins, right=thins)
    if len(', '.join(dannye_payment['chedp']) + ' ' + dannye_payment['dte_ched']) > 90:
        ws.row_dimensions[23].height = ceil(
            len(', '.join(dannye_payment['chedp']) + ' ' + dannye_payment['dte_ched']) / 100) * 14.4
    ws.merge_cells('B23:M23')

    top_left_cell = ws['B24']
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.font = Font(size=8)
    top_left_cell.border = Border(left=thins, right=thins, bottom=thins)
    top_left_cell.value = '(numeris, data)'
    ws.merge_cells('B24:M24')

    top_left_cell = ws['B25']
    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.font = Font(size=8)
    top_left_cell.border = Border(left=thins, right=thins)
    top_left_cell.value = '7. Valstybės rinkliavos suma:'
    ws.merge_cells('B25:M25')

    top_left_cell = ws['B26']
    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.font = Font(size=8, underline='single')
    top_left_cell.border = Border(left=thins, right=thins)
    top_left_cell.value = '__________________________________________________________________EUR'
    ws.merge_cells('B26:M26')

    top_left_cell = ws['B27']
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.font = Font(size=8)
    top_left_cell.border = Border(left=thins, right=thins, bottom=thins)
    top_left_cell.value = '(nurodyti sumą skaičiais ir žodžiais)'
    ws.merge_cells('B27:M27')

    top_left_cell = ws['B28']
    top_left_cell.alignment = Alignment(horizontal='left')
    top_left_cell.font = Font(size=8)
    top_left_cell.border = Border(left=thins, right=thins)
    top_left_cell.value = '8. Pasienio veterinarijos posto  pareigūnas: '
    ws.merge_cells('B28:M28')

    top_left_cell = ws['B29']
    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.border = Border(left=thins)
    top_left_cell.value = '_______________________'
    ws.merge_cells('B29:E29')

    top_left_cell = ws['F29']
    top_left_cell.alignment = Alignment(horizontal='center')
    top_left_cell.value = '_______________________'
    ws.merge_cells('F29:I29')

    ws['M29'].border = Border(right=thins)

    top_left_cell = ws['B30']
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.border = Border(left=thins)
    top_left_cell.font = Font(size=8)
    top_left_cell.value = '(vardas, pavardė)'
    ws.merge_cells('B30:E30')

    top_left_cell = ws['F30']
    top_left_cell.alignment = Alignment(horizontal='center', vertical='top')
    top_left_cell.font = Font(size=8)
    top_left_cell.value = '(parašas)'
    ws.merge_cells('F30:I30')

    ws['M30'].border = Border(right=thins)
    ws['B31'].border = Border(left=thins)
    ws['M31'].border = Border(right=thins)

    ws['B32'].border = Border(left=thins, bottom=thins)
    ws['B32'].value = 'Antspaudas'
    ws['B32'].font = Font(size=8)
    ws['C32'].border = Border(bottom=thins)
    ws['D32'].border = Border(bottom=thins)
    ws['E32'].border = Border(bottom=thins)
    ws['F32'].border = Border(bottom=thins)
    ws['G32'].border = Border(bottom=thins)
    ws['H32'].border = Border(bottom=thins)
    ws['I32'].border = Border(bottom=thins)
    ws['J32'].border = Border(bottom=thins)
    ws['K32'].border = Border(bottom=thins)
    ws['L32'].border = Border(bottom=thins)
    ws['M32'].border = Border(bottom=thins, right=thins)

    ws.page_setup.scale = 75
    folder_path = 'D:\\td'
    #name = 'D:\\td'

    file_name = dannye_payment['platezhka'] + '.xlsx'
    file_path = os.path.join(folder_path, file_name)


    wb.save(file_path)


    pythoncom.CoInitialize()
    Excel = win32com.client.Dispatch("Excel.Application")
    Excel.Visible = 0
    wb1 = Excel.Workbooks.Open(file_path)

    file_name = dannye_payment['platezhka']
    file_path = os.path.join(folder_path, file_name)

    wb1.ExportAsFixedFormat(0, file_path + '.pdf')
    wb1.Close()
